"""
Generate a formatted Excel workbook from detected no-solar warehouse data.
Two sheets:
  1. Warehouses     — one row per site with address, area, cost & solar estimates
  2. Summary        — portfolio totals and key metrics
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, Reference

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
OUT_PATH = DATA_DIR / "warehouses_no_solar.xlsx"

# ── Palette ────────────────────────────────────────────────────────────
NAVY   = "1F3864"
ORANGE = "C55A11"
AMBER  = "F4B942"
LIGHT  = "D9E1F2"
WHITE  = "FFFFFF"
GREEN  = "375623"
RED    = "C00000"
GREY   = "F2F2F2"


def _thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_font(size=11, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _cell_font(size=10, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def enrich(df: pd.DataFrame) -> pd.DataFrame:
    """Add cost-estimate columns to the warehouse DataFrame."""
    from src.analysis.baseline_demand import BaselineDemandAnalyser
    from src.analysis.peak_offpeak import PeakOffPeakAnalyser
    from src.cost_estimation.tariff_calculator import TariffCalculator

    rows = []
    for _, row in df.iterrows():
        ba = BaselineDemandAnalyser(warehouse_area_sqm=row["estimated_area_sqm"])
        profile = ba.benchmark_interval_kw()
        pa = PeakOffPeakAnalyser(profile)
        calc = TariffCalculator(pa, roof_area_sqm=row["estimated_area_sqm"])
        r = calc.annual_cost_report()
        rows.append({
            "annual_kwh": r.total_kwh,
            "peak_kwh": r.peak_kwh,
            "offpeak_kwh": r.offpeak_kwh,
            "annual_cost_aud": r.total_annual_cost,
            "solar_kwp": r.installable_kwp or 0,
            "solar_annual_kwh": r.solar_annual_kwh or 0,
            "solar_saving_aud": r.solar_annual_saving or 0,
            "solar_capex_aud": r.solar_capex or 0,
            "payback_years": r.solar_simple_payback_years or 0,
        })
    return pd.concat([df.reset_index(drop=True), pd.DataFrame(rows)], axis=1)


def write_warehouses_sheet(ws, df: pd.DataFrame) -> None:
    # ── Title bar ───────────────────────────────────────────────────────
    ws.merge_cells("A1:O1")
    title = ws["A1"]
    title.value = "Greater Sydney — Warehouses With No Solar Generation"
    title.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:O2")
    sub = ws["A2"]
    sub.value = "Buildings >500 m² · Ausgrid EA305 tariff · Sydney, NSW · 2024-25"
    sub.font = Font(name="Calibri", size=9, italic=True, color="595959")
    sub.fill = PatternFill("solid", fgColor=LIGHT)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Column headers ──────────────────────────────────────────────────
    headers = [
        ("#",           5),
        ("OSM ID",      12),
        ("Name",        22),
        ("Address",     35),
        ("Suburb",      18),
        ("Lat",         11),
        ("Lng",         11),
        ("Roof Area\n(m²)", 12),
        ("Solar\nDetected", 10),
        ("Annual\nkWh", 11),
        ("Annual Bill\n(AUD $)", 13),
        ("Solar\nCapacity (kWp)", 14),
        ("Solar\nSaving ($/yr)", 14),
        ("Solar\nCapex ($)", 13),
        ("Payback\n(Years)", 11),
    ]
    col_widths = {}
    for col_idx, (label, width) in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = _header_font(size=9)
        cell.fill = PatternFill("solid", fgColor=ORANGE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
        col_widths[col_idx] = width
    ws.row_dimensions[3].height = 30

    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Data rows ───────────────────────────────────────────────────────
    for row_idx, (_, row) in enumerate(df.iterrows(), start=4):
        shade = GREY if row_idx % 2 == 0 else WHITE
        fill = PatternFill("solid", fgColor=shade)

        values = [
            row_idx - 3,
            row["osm_id"],
            row.get("name", "") or "",
            row.get("address", "") or "",
            row.get("suburb", "") or "",
            round(float(row["lat"]), 6),
            round(float(row["lng"]), 6),
            round(float(row["estimated_area_sqm"]), 1),
            "No",
            round(float(row["annual_kwh"]), 0),
            round(float(row["annual_cost_aud"]), 0),
            round(float(row["solar_kwp"]), 1),
            round(float(row["solar_saving_aud"]), 0),
            round(float(row["solar_capex_aud"]), 0),
            round(float(row["payback_years"]), 1),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = _thin_border()
            cell.font = _cell_font()
            cell.alignment = Alignment(horizontal="right" if col_idx > 5 else "left",
                                       vertical="center")

        # Highlight "No" in red
        no_cell = ws.cell(row=row_idx, column=9)
        no_cell.font = Font(name="Calibri", size=10, bold=True, color=RED)
        no_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Currency format for cost columns
        for col in (11, 13, 14):
            ws.cell(row=row_idx, column=col).number_format = '$#,##0'
        ws.cell(row=row_idx, column=10).number_format = '#,##0'
        ws.cell(row=row_idx, column=8).number_format = '#,##0'

    # Freeze panes below header
    ws.freeze_panes = "A4"

    # Colour-scale on Annual Bill column
    last_row = 3 + len(df)
    bill_col = "K"
    ws.conditional_formatting.add(
        f"{bill_col}4:{bill_col}{last_row}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color=AMBER,
            end_type="max", end_color="F8696B",
        ),
    )


def write_summary_sheet(ws, df: pd.DataFrame) -> None:
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22

    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = "Portfolio Summary"
    t.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    metrics = [
        ("Total warehouses surveyed",          len(df)),
        ("Total roof area (m²)",               f"{df['estimated_area_sqm'].sum():,.0f}"),
        ("Largest site (m²)",                  f"{df['estimated_area_sqm'].max():,.0f}"),
        ("Smallest site (m²)",                 f"{df['estimated_area_sqm'].min():,.0f}"),
        ("", ""),
        ("── Energy & Cost ──",                ""),
        ("Combined annual consumption (MWh)",  f"{df['annual_kwh'].sum()/1000:,.0f}"),
        ("Combined annual electricity bill",   f"${df['annual_cost_aud'].sum():,.0f}"),
        ("Average bill per site",              f"${df['annual_cost_aud'].mean():,.0f}"),
        ("", ""),
        ("── Solar Opportunity ──",            ""),
        ("Total installable solar (kWp)",      f"{df['solar_kwp'].sum():,.0f}"),
        ("Total solar generation (MWh/yr)",    f"{df['solar_annual_kwh'].sum()/1000:,.0f}"),
        ("Total annual saving if solar added", f"${df['solar_saving_aud'].sum():,.0f}"),
        ("Total capital cost (solar)",         f"${df['solar_capex_aud'].sum():,.0f}"),
        ("Average simple payback",             f"{df['payback_years'].mean():.1f} years"),
    ]

    for i, (label, value) in enumerate(metrics, start=2):
        lc = ws.cell(row=i, column=1, value=label)
        vc = ws.cell(row=i, column=2, value=value)
        bg = GREY if i % 2 == 0 else WHITE
        for c in (lc, vc):
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = _thin_border()
            c.font = _cell_font(bold=(label.startswith("──")))
            c.alignment = Alignment(vertical="center",
                                    horizontal="right" if c.column == 2 else "left")
        ws.row_dimensions[i].height = 18
        if label.startswith("──"):
            lc.font = Font(name="Calibri", size=9, bold=True, color="595959")
            lc.fill = PatternFill("solid", fgColor=LIGHT)
            vc.fill = PatternFill("solid", fgColor=LIGHT)

    # ── Bar chart: Annual Bill by site ──────────────────────────────────
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Annual Electricity Bill by Warehouse (AUD)"
    chart.y_axis.title = "Site"
    chart.x_axis.title = "AUD $"
    chart.style = 10
    chart.width = 18
    chart.height = 12

    data_ref = Reference(
        ws.parent["Warehouses"],
        min_col=11, max_col=11,
        min_row=3, max_row=3 + len(df),
    )
    cats_ref = Reference(
        ws.parent["Warehouses"],
        min_col=2, max_col=2,
        min_row=4, max_row=3 + len(df),
    )
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "D2")


if __name__ == "__main__":
    df = pd.read_csv(DATA_DIR / "detected_warehouses.csv")
    df["address"] = df["address"].astype(str).replace("nan", "")
    df["name"] = df["name"].astype(str).replace("nan", "")
    df["suburb"] = df["suburb"].astype(str).replace("nan", "")

    print("Calculating cost estimates…")
    df = enrich(df)

    wb = Workbook()
    ws_wh = wb.active
    ws_wh.title = "Warehouses"
    ws_summary = wb.create_sheet("Summary")

    print("Writing Warehouses sheet…")
    write_warehouses_sheet(ws_wh, df)

    print("Writing Summary sheet…")
    write_summary_sheet(ws_summary, df)

    wb.save(str(OUT_PATH))
    print(f"\nSaved → {OUT_PATH}")
